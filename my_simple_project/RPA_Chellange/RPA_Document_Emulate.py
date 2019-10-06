from selenium import webdriver
import openpyxl

# def szybko(): # funkcja do podmiany nazw kolumn żebt odpowiadały nazwom labelek z rpa chellange
#     for idx, value in enumerate(col_names):
#         if value in "Phone Number":
#             col_names[idx]="Phone"
#         if value in "Role in Company":
#             col_names[idx]="Role"
#



driver = webdriver.Chrome()
driver.get("http://rpachallenge.com")
driver.maximize_window()

Wb = openpyxl.load_workbook('challenge.xlsx') # pobranie
Ws = Wb.active
col_names = [Ws.cell(row=1, column=j).value.strip() for j in range(1, 8)]
startbtn = driver.find_element_by_xpath("//button[@class='waves-effect col s12 m12 l12 btn-large uiColorButton']")
driver.execute_script("arguments[0].click();",startbtn)
vals=[]
#szybko()
for i in range(2, 12):
    vals = [Ws.cell(row=i, column=j).value for j in range(1, 8)]
    script =["document.evaluate('//label[text()="+'"'+str(col_names[idx])+'"]'+"'"+",document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue.parentElement.getElementsByTagName("+'"'+"input"+'")[0].value="'+str(val)+'"' for idx, val in enumerate(vals)] # druga metoda JS do uzupełnienia RPA chellange
    print(type(script))
    script_executed = ";".join(script) # średnika do każdego elementu z listy
    driver.execute_script(script_executed)
    driver.execute_script("document.getElementsByClassName('btn uiColorButton')[0].click()")
