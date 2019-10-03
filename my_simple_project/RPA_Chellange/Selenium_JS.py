from selenium import webdriver

from openpyxl import load_workbook

import openpyxl

driver = webdriver.Chrome()
driver.get("http://rpachallenge.com")
driver.maximize_window()

Wb = openpyxl.load_workbook('challenge.xlsx')
Ws = Wb.active
col_names = [Ws.cell(row=1, column=j).value.strip() for j in range(1, 8)]

driver.execute_script("document.getElementById('start').click()")

for i in range(2, 12):
vals = [Ws.cell(row=i, column=j).value for j in range(1, 8)]
scripts = ["document.evaluate('//label[contains(text()," + '"' + col_names[
idx] + '"' + ")]', document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue.parentElement.getElementsByTagName('input')[0].setAttribute('value', '" + str(
val) + "')" for idx, val in enumerate(vals)]
script_executed = ";".join(scripts)
driver.execute_script(script_executed)
driver.execute_script("document.getElementsByClassName('btn btn-default')[0].click()")
