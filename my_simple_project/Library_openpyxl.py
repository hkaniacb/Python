import openpyxl
from datetime import date
import os.path as op

name_workbook = 'sample.xlsx'
name_sheet = 'Ceneo'

print("Czy istnieje pilk: "+op.exists(name_workbook))

if op.exists(name_workbook):
    print('Exist workbook ' + name_workbook)
    wb = openpyxl.load_workbook(name_workbook)#   Workbook("sample.xlsx")
else:
    print('Create workbook ' + name_book)
    wb = openpyxl.Workbook()

#Sprawdza czy istanieje arkusz i tworzy jeśli nei
if name_sheet in wb.sheetnames:
    print('Exist sheet ' + name_sheet)
    ws =  wb[name_sheet]
else:
    print('Create Sheet '+name_sheet)
    ws = wb.create_sheet(name_sheet, 0)

#Dodaje wiersz
ws.append([date.today(),"Hello"])

#Zapisuje Workbook
wb.save(name_workbook)




#Zapis do  Exceli przy użyciu Classy 


import openpyxl
from datetime import date
import os.path as op


class ExcelPrice:

    def __init__(self, path_excel):  # Konstruktor sprawdza czy insteje workbook o podanej lokalizacji jeśli nie ma to go tworzy
        if op.exists(path_excel): # warunek sprawdzjący czy istnieje pilk
            print('Exist workbook ' + path_excel)
            self.wb = openpyxl.load_workbook(path_excel)  # Załaduj istniejący plik
        else:
            print('Create workbook ')
            self.wb =  openpyxl.Workbook()

    def check_exist_sheet(self, s): # Sprawdza czy istnieje arkusz dla danego workbooka
        # Sprawdza czy istanieje arkusz i tworzy jeśli nei
        if s in self.wb.sheetnames:
            print('Exist sheet ' + s)
            self.ws = self.wb[s]
        else:
            print('Create Sheet ' + s)
            self.ws = self.wb.create_sheet(s, 0)

    def append_row(self, lista_parameters):
        # Dodaje wiersz
        temp_param = [date.today()]
        for item in lista_parameters:
            temp_param.append(item)

        self.ws.append(temp_param) #Dodanie rekordu do excela
        self.wb.save(name_workbook) # Zapisuje Workbook


name_workbook = 'sample3.xlsx'
name_sheet = 'Ceneo3'
myExel = ExcelPrice(name_workbook)
myExel.check_exist_sheet(name_sheet)

my_list = ['geeks', 'for', 'geeks'] #test
another_list = "dad" #test

my_list.append(another_list) # test
myExel.append_row(my_list)

## czytanie z Exceli bez Classy 


# Python program to read an excel file
# potrzeba jeszcze biblioteka  openpyxl

# Give the location of the file
path = r"dataInput\DataInput.xlsx"

# To open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook(path)

# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active

counter = 1  # Licznik do chodzenia po wierszu 
while True:

    cell_obj = sheet_obj.cell(row=counter, column=1)

    # Print value of cell object
    # using the value attribute
    if cell_obj.value == None:
        break

    print(cell_obj.value)
    counter +=1
