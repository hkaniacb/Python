#Pobieranie wartości z Excela i stowrzenie słwonika podzielenie go na zdania
# Podzielenie Excela na zadania przy użyciu słwnika
from openpyxl import Workbook,load_workbook
import  datetime


def readExcelFile(path,slownik):
    wb = load_workbook(path, data_only=True) #  Tworzenie workbooka, (data_only pobierania komórki tylko z wartosciami)
    ws = wb.worksheets[0] # Pobieranie pierwszego arkusza z workbooka
    newdict = {} # Tworzenie nowego słownika
    for varKey in slownik["Pola kampanii"]:  # iteracja po słowniku kluczy ( kolumn w Exelu ) jako parametr fukncji
        for item in ws.iter_cols(): # iteracja po arkuszu w celu wyszukania odpowiednije kolumny  ze słwonika
            for indexcel, cell in enumerate(item): # załadowanie słownika wartościami z excela
                if str(varKey.upper()) == str(cell.value):
                    newlist = []
                    licznik = indexcel + 1
                    val = item[licznik].value
                    while str(val) != "None":
                        if isinstance(val, datetime.datetime): # sprawdzenie czy ten typ wartosci jest datą
                            newlist.append(val.strftime("%d-%m-%Y")) # formatowanie daty, do podanego schematu
                        else:
                            newlist.append(val) # dodawnie do listy komórek  z excela
                        licznik += 1
                        val = item[licznik].value
                    newdict[str(varKey.upper())] = newlist
    return newdict

def listaZadan(sKolumn): # Parsowanie na zadania w słowniku
 listaZadan=[]
 for i in range(len(max(sKolumn.values()))): # iteracja po maksymalnej wartosci w liscie 
    newdict={} # tworzenie docelowego słownika, który będzie wykorzystany w projecie 
    for item in Dictionary .keys():
        try:
         newdict[item]=sKolumn[item][i]
        except:
            newdict[item] = "None"
            continue
    listaZadan.append(newdict)
 return listaZadan

pathExcel=r"D:\baza.xlsx" # przypisanie surogwego stringa przy użyciu przedtrostka "r"
cExcelFiledColumn={"Pola kampanii":["Kampania","TYP EMISJI","Kontraktor","Witryna","forma reklamy","rozmiar reklamy","limit wyœwietleñ","start","koniec"]} # Config wejsciowy z kolumnami 
Dictionary =readExcelFile(pathExcel,cExcelFiledColumn)  # wywowanie funkcji pobieracaej Excela
Dictionary =listaZadan(Dictionary )
for index,value in enumerate(Dictionary ):
    print(str(index)+ " " + str(value))
